import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from compare_support import merge_access_values, parse_cdc_sheet, normalize_door_code


class CompareSupportTests(unittest.TestCase):
    def test_merge_duplicate_columns(self):
        door_columns = {"L01": [3, 4]}
        row_values = {3: "", 4: "X"}

        def lookup(col):
            return row_values.get(col)

        access = merge_access_values(door_columns, lookup)
        self.assertTrue(access["L01"])

    def test_parse_cdc_sheet_with_duplicates(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "CDC"

        ws.cell(row=1, column=1, value="Service")
        ws.cell(row=1, column=2, value="Function")
        ws.cell(row=1, column=3, value="L01")
        ws.cell(row=1, column=4, value="L01")
        ws.cell(row=1, column=5, value="L02")

        ws.cell(row=2, column=1, value="AQ")
        ws.cell(row=2, column=2, value="Func A")
        ws.cell(row=2, column=4, value="X")

        ws.cell(row=3, column=1, value="AQ")
        ws.cell(row=3, column=2, value="Func B")
        ws.cell(row=3, column=5, value="x")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
            temp_path = Path(handle.name)
        try:
            wb.save(temp_path)
            data = parse_cdc_sheet(temp_path, sheet_name="CDC")
            self.assertEqual(data["doors"], [normalize_door_code("L01"), normalize_door_code("L02")])

            func_a = data["functions"][0]
            func_b = data["functions"][1]
            self.assertTrue(func_a["access"]["L01"])
            self.assertFalse(func_a["access"]["L02"])
            self.assertTrue(func_b["access"]["L02"])
        finally:
            temp_path.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
