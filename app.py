from __future__ import annotations

import csv
import json
import os
from datetime import datetime
from difflib import SequenceMatcher
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

# Helper functions for normalization (moved from compare_support)
def normalize_door_code(code: str) -> str:
    if not code:
        return ""
    code = str(code).strip().upper()
    if code.startswith("L0"):
        return "L" + code[2:]
    return code

def normalize_label(label: str) -> str:
    import re
    if not label:
        return ""
    # Remove special chars and lowercase
    return re.sub(r'[^a-zA-Z0-9]', '', str(label)).lower()

BASE_DIR = Path(__file__).resolve().parent
WEB_DIR = BASE_DIR / "web"
LOCAL_DATA_DIR = BASE_DIR / "data"
DEFAULT_DOORS_PATH = (
    str(LOCAL_DATA_DIR / "doors")
    if (LOCAL_DATA_DIR / "doors").is_dir()
    else "\\\\files\\partage\\Donn\u00e9es Cytopharma\\suprema\\doors"
)
DEFAULT_SUPREMA_PATH = (
    str(LOCAL_DATA_DIR / "support" / "suprema.xlsx")
    if (LOCAL_DATA_DIR / "support" / "suprema.xlsx").is_file()
    else "\\\\files\\partage\\Donn\u00e9es Cytopharma\\suprema\\support\\suprema.xlsx"
)
DEFAULT_SUPPORT_JSON_PATH = BASE_DIR / "exports" / "doors_reference.json"


def resolve_config_path(value: str | None, default: str | Path) -> str:
    raw = value if value not in (None, "") else default
    path = Path(raw)
    if not path.is_absolute():
        path = BASE_DIR / path
    return str(path)


DOORS_PATH = resolve_config_path(os.environ.get("DOORS_PATH"), DEFAULT_DOORS_PATH)
SUPREMA_PATH = resolve_config_path(os.environ.get("SUPREMA_PATH"), DEFAULT_SUPREMA_PATH)
SUPPORT_JSON_PATH = resolve_config_path(
    os.environ.get("SUPPORT_JSON_PATH"), DEFAULT_SUPPORT_JSON_PATH
)
CDC_SHEET = os.environ.get("CDC_SHEET", "CDC")
DEPARTMENTS_FILE = os.environ.get("DEPARTMENTS_FILE", "")
DOOR_METADATA_FILE = os.environ.get("DOOR_METADATA_FILE", "")
HOST = os.environ.get("HOST", "127.0.0.1")
PORT = int(os.environ.get("PORT", "8000"))

REFERENCE_DOORS = [
    {
        "code": "L01",
        "label": "Admin-In",
        "group": "Administration",
        "aliases": ["Administra IN", "Admin In", "Admin-In"],
    },
    {
        "code": "L02",
        "label": "Admin-Out",
        "group": "Administration",
        "aliases": ["Administra OUT", "Admin Out", "Admin-Out"],
    },
    {
        "code": "L03",
        "label": "Vest PROD IN",
        "group": "PROD",
        "aliases": ["VEST PROD IN", "Vest PROD IN"],
    },
    {
        "code": "L04",
        "label": "Vest PROD OUT",
        "group": "PROD",
        "aliases": ["VEST PROD OUT", "Vest PROD OUT"],
    },
    {
        "code": "L05",
        "label": "Labo CQ In",
        "group": "Labo CQ",
        "aliases": ["LABO IN", "Labo IN", "Labo CQ IN", "Labo CQ In"],
    },
    {
        "code": "L06",
        "label": "Labo CQ Out",
        "group": "Labo CQ",
        "aliases": ["LABO OUT", "Labo OUT", "Labo CQ OUT", "Labo CQ Out"],
    },
    {
        "code": "L07",
        "label": "MAG In",
        "group": "MAG",
        "aliases": ["MAG IN", "MAG In"],
    },
    {
        "code": "L08",
        "label": "MAG Out",
        "group": "MAG",
        "aliases": ["MAG OUT", "MAG Out"],
    },
    {
        "code": "L09",
        "label": "MAG 2 / MAG IN MP",
        "group": "MAG",
        "aliases": ["MAG 2", "MAG IN MP", "MAG2"],
    },
    {
        "code": "L11",
        "label": "Salle serveur",
        "group": "Serveur",
        "aliases": ["IN SAL SERVEUR", "SAL SERVEUR", "SALLE SERVEUR"],
    },
    {
        "code": "L14",
        "label": "SAS VIS 1 IN",
        "group": "VIS",
        "aliases": ["SAS VIS 1 IN", "SasV1In", "SASV1IN"],
    },
    {
        "code": "L15",
        "label": "SAS VIS 2 IN",
        "group": "VIS",
        "aliases": ["SAS VIS 2 IN", "SasV2In", "SASV2IN"],
    },
    {
        "code": "L16",
        "label": "OUT SAS MP MAG",
        "group": "MAG",
        "aliases": ["OUT SAS MP MAG"],
    },
    {
        "code": "L17",
        "label": "OUT SAS MP PROD",
        "group": "PROD",
        "aliases": ["OUT SAS MP PROD"],
    },
    {
        "code": "L18",
        "label": "OUT SAS PF MAG",
        "group": "MAG",
        "aliases": ["OUT SAS PF MAG"],
    },
    {
        "code": "L19",
        "label": "OUT SAS PF PROD",
        "group": "PROD",
        "aliases": ["OUT SAS PF PROD"],
    },
]


def parse_int(value):
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def parse_float(value):
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def find_case_insensitive(directory: Path, filename: str) -> Path | None:
    target = filename.lower()
    for path in directory.iterdir():
        if path.name.lower() == target:
            return path
    return None


def read_csv_rows(path: Path):
    with path.open("r", newline="", encoding="utf-8-sig", errors="replace") as handle:
        return list(csv.reader(handle))


def load_users(doors_path: Path):
    users_file = find_case_insensitive(doors_path, "users.csv")
    if not users_file:
        raise FileNotFoundError("users.csv not found in doors path")

    rows = read_csv_rows(users_file)
    users = []
    for row in rows:
        if not row:
            continue
        user_id = parse_int(row[0])
        if user_id is None:
            continue
        name = row[1].strip() if len(row) > 1 else ""
        function = row[2].strip() if len(row) > 2 else ""
        users.append({"id": user_id, "name": name, "function": function})

    return users, users_file


def resolve_suprema_path():
    path = Path(SUPREMA_PATH)
    if path.is_file():
        return path
    local = BASE_DIR / "suprema.xlsx"
    if local.is_file():
        return local
    return None


def load_department_map_from_csv(map_path: Path):
    rows = read_csv_rows(map_path)
    if not rows:
        return {}, map_path

    header = [cell.strip().lower() for cell in rows[0]]
    if "function" in header and "department" in header:
        idx_function = header.index("function")
        idx_department = header.index("department")
        start = 1
    else:
        idx_function = 0
        idx_department = 1
        start = 0

    mapping = {}
    for row in rows[start:]:
        if len(row) <= max(idx_function, idx_department):
            continue
        function = row[idx_function].strip()
        department = row[idx_department].strip()
        if function and function not in mapping:
            mapping[function] = department

    return mapping, map_path


def load_department_map_from_suprema(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read suprema.xlsx") from exc

    wb = load_workbook(path, data_only=True, read_only=True)
    if CDC_SHEET not in wb.sheetnames:
        raise ValueError(f"sheet '{CDC_SHEET}' not found in {path}")
    ws = wb[CDC_SHEET]

    header_row = None
    for index, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        if not row or len(row) < 2:
            continue
        left = str(row[0]).strip().lower() if row[0] is not None else ""
        right = str(row[1]).strip().lower() if row[1] is not None else ""
        if left == "service" and right.startswith("fonct"):
            header_row = index
            break

    start_row = header_row + 1 if header_row else 1
    mapping = {}
    current_service = ""
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if not row or len(row) < 2:
            continue
        service = str(row[0]).strip() if row[0] is not None else ""
        function = str(row[1]).strip() if row[1] is not None else ""
        if service:
            current_service = service
        if function and function not in mapping:
            mapping[function] = current_service

    return mapping, path


def load_department_map(doors_path: Path):
    if DEPARTMENTS_FILE:
        map_path = Path(DEPARTMENTS_FILE)
        if not map_path.is_file():
            raise FileNotFoundError(f"departments file not found: {map_path}")
        return load_department_map_from_csv(map_path)

    suprema_path = resolve_suprema_path()
    if suprema_path:
        return load_department_map_from_suprema(suprema_path)

    map_path = find_case_insensitive(doors_path, "departments.csv")
    if not map_path:
        return {}, None

    return load_department_map_from_csv(map_path)


def is_access_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().upper()
    return text in {"X", "1", "YES", "Y", "TRUE"}


def guess_door_group(label: str) -> str:
    name = label.lower()
    if "admin" in name:
        return "Administration"
    if "serveur" in name or "server" in name:
        return "Serveur"
    if "labo" in name:
        return "Labo CQ"
    if "sas" in name or "vis" in name:
        return "VIS"
    if "prod" in name:
        return "PROD"
    if "mag" in name:
        return "MAG"
    return ""


def load_door_metadata(doors_path: Path):
    if DOOR_METADATA_FILE:
        meta_path = Path(DOOR_METADATA_FILE)
        if not meta_path.is_file():
            raise FileNotFoundError(f"door metadata file not found: {meta_path}")
    else:
        meta_path = find_case_insensitive(doors_path, "door_metadata.csv")
        if not meta_path:
            return {}, None

    rows = read_csv_rows(meta_path)
    if not rows:
        return {}, meta_path

    header = [cell.strip().lower() for cell in rows[0]]
    known = {
        "file",
        "door_file",
        "filename",
        "name",
        "label",
        "door_label",
        "display",
        "group",
        "door_group",
        "department",
        "code",
        "door_code",
        "order",
        "sort",
        "position",
    }
    has_header = any(cell in known for cell in header)

    if has_header:
        start = 1
        idx_file = next((i for i, cell in enumerate(header) if cell in {"file", "door_file", "filename", "name"}), None)
        idx_label = next((i for i, cell in enumerate(header) if cell in {"label", "door_label", "display"}), None)
        idx_group = next((i for i, cell in enumerate(header) if cell in {"group", "door_group", "department"}), None)
        idx_code = next((i for i, cell in enumerate(header) if cell in {"code", "door_code"}), None)
        idx_order = next((i for i, cell in enumerate(header) if cell in {"order", "sort", "position"}), None)
    else:
        start = 0
        idx_file, idx_label, idx_group, idx_code, idx_order = 0, 1, 2, 3, 4

    metadata = {}
    for row in rows[start:]:
        if idx_file is None or len(row) <= idx_file:
            continue
        file_name = row[idx_file].strip()
        if not file_name:
            continue
        label = row[idx_label].strip() if idx_label is not None and len(row) > idx_label else ""
        group = row[idx_group].strip() if idx_group is not None and len(row) > idx_group else ""
        code = row[idx_code].strip() if idx_code is not None and len(row) > idx_code else ""
        order = None
        if idx_order is not None and len(row) > idx_order:
            order = parse_float(row[idx_order])
        metadata[file_name.lower()] = {
            "label": label,
            "group": group,
            "code": code,
            "order": order,
        }

    return metadata, meta_path


def build_reference_alias_map():
    alias_map = {}
    for ref in REFERENCE_DOORS:
        aliases = [ref["code"], ref["label"]] + ref.get("aliases", [])
        for alias in aliases:
            key = normalize_label(alias)
            if not key:
                continue
            alias_map.setdefault(key, ref["code"])
    return alias_map


def match_reference_code(label: str, alias_map: dict[str, str]) -> str:
    if not label:
        return ""
    direct = normalize_door_code(label)
    if direct and direct.startswith("L") and direct[1:].isdigit():
        return direct

    key = normalize_label(label)
    if key in alias_map:
        return alias_map[key]

    best_code = ""
    best_score = 0.0
    for candidate, code in alias_map.items():
        score = SequenceMatcher(None, key, candidate).ratio()
        if score > best_score:
            best_score = score
            best_code = code
    if best_score >= 0.6:
        return best_code
    return ""



_DASHBOARD_CACHE = {"data": None, "expiry": 0}


def get_cached_dashboard(ttl: int = 5):
    now = datetime.now().timestamp()
    if _DASHBOARD_CACHE["data"] and now < _DASHBOARD_CACHE["expiry"]:
        return _DASHBOARD_CACHE["data"]

    data = build_dashboard()
    _DASHBOARD_CACHE["data"] = data
    _DASHBOARD_CACHE["expiry"] = now + ttl
    return data


def load_doors(doors_path: Path):
    metadata, metadata_path = load_door_metadata(doors_path)
    alias_map = build_reference_alias_map()
    reference_codes = {ref["code"] for ref in REFERENCE_DOORS}
    door_files = []
    for path in sorted(doors_path.glob("*.csv"), key=lambda p: p.name.lower()):
        name_lower = path.name.lower()
        if name_lower in ("users.csv", "departments.csv", "door_metadata.csv"):
            continue
        door_files.append(path)

    code_to_users = {code: set() for code in reference_codes}

    for path in door_files:
        base_name = path.stem.strip() or path.stem
        meta = metadata.get(path.name.lower()) or metadata.get(base_name.lower()) or {}
        code = meta.get("code") or ""
        code = normalize_door_code(code) if code else ""
        if code and code not in reference_codes:
            code = ""
        if not code:
            code = match_reference_code(base_name, alias_map)
        if not code or code not in reference_codes:
            continue

        ids = set()
        for row in read_csv_rows(path):
            if not row:
                continue
            user_id = parse_int(row[0])
            if user_id is None:
                continue
            ids.add(user_id)

        code_to_users.setdefault(code, set()).update(ids)

    door_names = []
    door_meta = []
    door_to_users = {}

    for ref in REFERENCE_DOORS:
        code = ref["code"]
        label = ref["label"]
        group = ref.get("group") or guess_door_group(label)
        display = f"{code} {label}".strip()
        door_names.append(display)
        door_to_users[display] = code_to_users.get(code, set())
        door_meta.append(
            {
                "label": label,
                "group": group,
                "code": code,
            }
        )

    return door_names, door_to_users, door_meta, door_files, metadata_path


def build_dashboard():
    doors_path = Path(DOORS_PATH)
    if not doors_path.exists():
        raise FileNotFoundError(f"doors path not found: {doors_path}")

    print("[DEBUG] Starting build_dashboard...")
    users, users_file = load_users(doors_path)
    print(f"[DEBUG] Loaded {len(users)} users from {users_file}")
    
    dept_map, dept_file = load_department_map(doors_path)
    print(f"[DEBUG] Loaded department map from {dept_file or 'None'}")
    
    print("[DEBUG] Loading doors (this might take a moment)...")
    door_names, door_to_users, door_meta, door_files, metadata_path = load_doors(doors_path)
    print(f"[DEBUG] Loaded {len(door_names)} doors")

    rows = []
    for user in users:
        department = dept_map.get(user["function"], "")
        row = [department, user["id"], user["name"], user["function"]]
        for door in door_names:
            row.append("X" if user["id"] in door_to_users.get(door, set()) else "")
        rows.append(row)

    return {
        "meta": {
            "doors_path": str(doors_path),
            "users_file": str(users_file),
            "departments_file": str(dept_file) if dept_file else "",
            "door_metadata_file": str(metadata_path) if metadata_path else "",
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "door_count": len(door_names),
            "user_count": len(users),
        },
        "columns": ["Department", "ID", "Name", "Function"] + door_names,
        "doors": door_names,
        "door_meta": door_meta,
        "rows": rows,
    }


def filter_rows(payload, query: str, door: str, only_access: bool):
    rows = payload["rows"]
    if query:
        needle = query.lower()
        rows = [row for row in rows if needle in " ".join(str(cell) for cell in row).lower()]

    if door and door != "all" and only_access:
        try:
            door_index = payload["columns"].index(door)
        except ValueError:
            return []
        rows = [row for row in rows if row[door_index] == "X"]

    return rows


def build_export_bytes(payload, rows):
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    door_meta = payload.get("door_meta") or [
        {"label": name, "group": "", "code": ""} for name in payload.get("doors", [])
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    header_rows = 4
    data_start = header_rows + 1
    user_columns = ["Department", "ID", "Name", "Function"]

    thin = Side(style="thin", color="BFBFBF")
    border = Border(top=thin, right=thin, bottom=thin, left=thin)

    fill_group = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_door = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    fill_code = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    fill_header = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
    fill_yes = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_no = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    font_header = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_door = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)

    total_columns = len(user_columns) + len(door_meta)

    widths = [18, 10, 28, 32] + [6] * len(door_meta)
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 120
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 22

    door_start_col = len(user_columns) + 1

    # Row 1: group header
    current_group = None
    group_start = door_start_col
    for offset, meta in enumerate(door_meta):
        col = door_start_col + offset
        group = meta.get("group", "")
        cell = ws.cell(row=1, column=col, value=group)
        cell.fill = fill_group
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border
        if current_group is None:
            current_group = group
            group_start = col
        if group != current_group:
            if current_group:
                ws.merge_cells(start_row=1, start_column=group_start, end_row=1, end_column=col - 1)
            current_group = group
            group_start = col

    if door_meta:
        last_col = door_start_col + len(door_meta) - 1
        if current_group:
            ws.merge_cells(start_row=1, start_column=group_start, end_row=1, end_column=last_col)

    # Row 2: door labels
    for offset, meta in enumerate(door_meta):
        col = door_start_col + offset
        cell = ws.cell(row=2, column=col, value=meta.get("label", ""))
        cell.fill = fill_door
        cell.font = font_header
        cell.alignment = align_door
        cell.border = border

    # Row 3: door codes
    for offset, meta in enumerate(door_meta):
        col = door_start_col + offset
        cell = ws.cell(row=3, column=col, value=meta.get("code", ""))
        cell.fill = fill_code
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border

    # Row 4: column headers
    for col, label in enumerate(user_columns, start=1):
        cell = ws.cell(row=4, column=col, value=label)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border

    for offset in range(len(door_meta)):
        col = door_start_col + offset
        cell = ws.cell(row=4, column=col, value="Access")
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border

    # Data rows
    for row_index, row in enumerate(rows, start=data_start):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            if col_index >= door_start_col:
                cell.alignment = align_center
                if value == "X":
                    cell.fill = fill_yes
                else:
                    cell.value = ""
                    cell.fill = fill_no
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = f"{get_column_letter(door_start_col)}{data_start}"
    ws.auto_filter.ref = f"A4:{get_column_letter(total_columns)}{data_start + len(rows) - 1}"

    ws.protection.sheet = True
    ws.protection.set_password("0000")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class DashboardHandler(BaseHTTPRequestHandler):
    server_version = "SupremaDashboard/1.0"

    def log_message(self, format, *args):
        # Override to ensure logs are always printed to stdout
        print(f"[{self.log_date_time_string()}] {format % args}")

    def handle_error(self, exc):
        # Gracefully handle connection resets which are common in browser refreshes
        if isinstance(exc, (ConnectionAbortedError, ConnectionResetError)):
            return
        print(f"[ERROR] {exc}")

    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            query = parse_qs(parsed.query)

            if path == "/":
                return self.serve_file("index.html", "text/html; charset=utf-8")
            if path == "/styles.css":
                return self.serve_file("styles.css", "text/css; charset=utf-8")
            if path == "/app.js":
                return self.serve_file("app.js", "application/javascript; charset=utf-8")
            if path == "/api/dashboard":
                return self.serve_dashboard()
            if path == "/api/export.xlsx":
                return self.serve_export(query)
            if path == "/api/health":
                return self.serve_json({"ok": True})

            self.send_error(404, "Not Found")
        except (ConnectionAbortedError, ConnectionResetError):
            pass
        except Exception as exc:
            self.handle_error(exc)

    def serve_file(self, filename: str, content_type: str):
        file_path = WEB_DIR / filename
        if not file_path.is_file():
            self.send_error(404, "Not Found")
            return
        data = file_path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        try:
            self.wfile.write(data)
        except (ConnectionAbortedError, ConnectionResetError):
            pass

    def serve_json(self, payload, status=200):
        data = json.dumps(payload, ensure_ascii=True).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        try:
            self.wfile.write(data)
        except (ConnectionAbortedError, ConnectionResetError):
            pass

    def serve_dashboard(self):
        try:
            payload = get_cached_dashboard()
        except Exception as exc:
            return self.serve_json({"error": str(exc)}, status=500)
        return self.serve_json(payload)

    def serve_export(self, query):
        try:
            payload = get_cached_dashboard()
            term = (query.get("q") or [""])[0]
            door = (query.get("door") or ["all"])[0]
            only_access = (query.get("only_access") or ["0"])[0] == "1"
            rows = filter_rows(payload, term, door, only_access)
            data = build_export_bytes(payload, rows)
        except Exception as exc:
            return self.serve_json({"error": str(exc)}, status=500)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"door_access_{stamp}.xlsx"
        self.send_response(200)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Disposition", f"attachment; filename={filename}")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        try:
            self.wfile.write(data)
        except (ConnectionAbortedError, ConnectionResetError):
            pass

    def do_POST(self):
        if self.path == "/api/upload":
            return self.handle_upload()
        self.send_error(404, "Not Found")

    def handle_upload(self):
        try:
            content_type = self.headers.get('Content-Type')
            if not content_type or 'multipart/form-data' not in content_type:
                return self.serve_json({"error": "Content-Type must be multipart/form-data"}, status=400)
            
            boundary = content_type.split("boundary=")[-1]
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                return self.serve_json({"error": "Empty upload"}, status=400)

            # In Python 3.13, we manually parse multipart
            body = self.rfile.read(content_length)
            parts = self.parse_multipart(body, boundary)
            
            if not parts:
                return self.serve_json({"error": "No files found in upload"}, status=400)

            # Validation and Saving
            errors = self.process_upload_parts(parts)
            if errors:
                return self.serve_json({"error": "\n".join(errors)}, status=400)

            # Clear cache
            global _DASHBOARD_CACHE
            _DASHBOARD_CACHE = {"data": None, "expiry": 0}

            return self.serve_json({"ok": True, "message": f"Successfully synced {len(parts)} files"})
        except Exception as exc:
            return self.serve_json({"error": str(exc)}, status=500)

    def parse_multipart(self, body, boundary):
        import re
        delimiter = b'--' + boundary.encode('ascii')
        parts = []
        for raw_part in body.split(delimiter):
            if not raw_part or raw_part == b'--\r\n' or raw_part == b'--':
                continue
            header_end = raw_part.find(b'\r\n\r\n')
            if header_end == -1:
                continue
            headers = raw_part[2:header_end].decode('utf-8', errors='replace')
            data = raw_part[header_end+4:]
            if data.endswith(b'\r\n'):
                data = data[:-2]
            
            match = re.search(r'filename="([^"]+)"', headers)
            if match:
                parts.append({'filename': match.group(1), 'data': data})
        return parts

    def process_upload_parts(self, parts):
        errors = []
        files_to_save = []
        
        # Determine the "Required Set" from existing files
        doors_dir = Path('data/doors')
        required_files = {} # lowercase -> original
        if doors_dir.exists():
            for p in doors_dir.glob('*.csv'):
                required_files[p.name.lower()] = p.name
        
        has_users = False
        uploaded_names = {part['filename'].lower() for part in parts}
        
        for part in parts:
            name = part['filename'].lower()
            if name == 'users.csv':
                # Validate users.csv structure
                try:
                    content = part['data'].decode('utf-8-sig', errors='replace')
                    lines = content.strip().split('\n')
                    if not lines:
                        errors.append("users.csv is empty")
                    else:
                        cols = list(csv.reader([lines[0]]))[0]
                        if len(cols) < 2:
                            errors.append("users.csv must have at least 2 columns (ID, Name)")
                except Exception as e:
                    errors.append(f"Invalid format for users.csv: {e}")
                has_users = True
                files_to_save.append(('data/doors/users.csv', part['data']))
            
            elif name == 'suprema.xlsx':
                files_to_save.append(('data/support/suprema.xlsx', part['data']))
            
            elif name.endswith('.csv'):
                # Door files
                files_to_save.append((f'data/doors/{part["filename"]}', part['data']))

        # Check for missing required files
        for req_lower, req_orig in required_files.items():
            if req_lower not in uploaded_names:
                errors.append(f"Missing required file: {req_orig}")

        if not has_users and 'users.csv' in required_files:
            # Already handled by loop above, but just in case
            pass

        if errors:
            return errors

        # If we got here, everything is valid. Proceed to save.
        os.makedirs('data/doors', exist_ok=True)
        os.makedirs('data/support', exist_ok=True)
        
        # Save files (overwriting)
        for path, data in files_to_save:
            with open(path, 'wb') as f:
                f.write(data)
        
        return None

def main():
    try:
        server = ThreadingHTTPServer((HOST, PORT), DashboardHandler)
        print(f"Suprema Dashboard started at http://{HOST}:{PORT}")
        print(f"Using DOORS_PATH: {DOORS_PATH}")
        print(f"Using SUPREMA_PATH: {SUPREMA_PATH}")
        print("Press Ctrl+C to stop.")
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping server...")
        server.server_close()
    except Exception as exc:
        print(f"\nFailed to start server: {exc}")


if __name__ == "__main__":
    main()
